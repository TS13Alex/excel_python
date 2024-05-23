'''генерирует листы excel для заполнения ПЕ3, СП, ТС'''
from json import load
from pathlib import WindowsPath
from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Border, Side, Alignment, Font

NROW = 57 #total row 5 mm in 585 mm


def setup_row(ws, total_list:int, nrow:int):
    '''setup row height'''
    all_rows = nrow * total_list #total row in all lists
    row_heigh_5 = 14.6 #height row 5 мм
    for i in range (1, all_rows + 1):
        ws.row_dimensions[i].height = row_heigh_5 #make all row in 5 мм

def setup_col(ws):
    '''setup collumns'''
    name_col = 'ABCDEFGHIJKLMNOPQRST' #number of collums
    #COLLUMNS_WIDTH = (5,7,7,6,4,4,19,15,10,4,23,33,10,5,5,5,5,10,10,10, ) in mm
    col_width = (2.50, 3.50, 3.50, 3.00, 2.00, 2.00, 9.80, 7.50, 5.00, 2.00, 11.80,
    16.80, 5.00, 2.50, 2.50, 2.50, 2.50, 5.00, 5.00, 5.00) #width collums in pt excel
    for (i, j) in zip(name_col,col_width):
        ws.column_dimensions[i].width = j

def setup_sheet(ws):
    '''setup sheet properties'''
    ws.page_setup.paperSize = 9 #format paper А4
    ws.page_setup.orientation = 'portrait' #orientation portret
    #margins
    left = 0.27 #0.7/inch
    right = 0.2
    bottom = 0.24 #0.6/inch
    top = 0.22
    ws.page_margins = PageMargins(left, right, top, bottom)

def make_merge_cells(ws, xy_coordinat:list):
    '''merge cell and add border'''
    bd = Side(style='thin', color='000000')
    r1, c1, r2, c2 = xy_coordinat
    if r1 != r2 or c1 != c2:
        for i in range(r1, r2+1):
            for j in range (c1, c2+1):
                ws.cell(row=i, column=j).border = Border(left=bd, top=bd, right=bd, bottom=bd)
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    else:
        cell = ws.cell(row=r1, column=c1)
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

def make_ramka(ws, col:list, st_r:int):
    '''make ramka'''
    start = st_r - 1
    for i in col:
        r1, c1, r2, c2 = i[:4]
        make_merge_cells(ws, [start+r1, c1, start+r2, c2])
        add_text_cell(ws, [start+r1, c1, start+r2, c2]+i[4:])

def make_document(wb, n_a_f:tuple, type_doc:str, countslist:int, data_dev:list):
    '''generate excel sheet'''
    drawby = n_a_f[0]
    checkby = n_a_f[1]
    normby = n_a_f[2]
    approvby  = n_a_f[3]
    inv_num  = n_a_f[4]
    gbnk = n_a_f[5]
    namedevice = n_a_f[6]
    match type_doc:
        case "PE":
            gbnk_suff = 'ПЭ3'
            write_data = write_pe_data
        case "SP":
            gbnk_suff = 'СП'
            write_data = write_sp_data
        case "TC":
            gbnk_suff = 'ТС'
            write_data = write_tc_data
    for i_sh in wb.sheetnames:
        wb.remove(wb[i_sh])
    wspe = wb.create_sheet(type_doc)
    setup_sheet(wspe)
    setup_col(wspe)
    setup_row(wspe, countslist, NROW)
    m = WindowsPath(__file__)
    t = m.with_name("sxema.json")
    with open(t, "r", encoding="utf-8") as fh:
        col = load(fh) # download structure from file
    #first list document
    start = 1
    col["LEFT_FIRST_FRAME"][1][4] = gbnk #add gbnk
    make_ramka(wspe, col["LEFT_FIRST_FRAME"], start)
    col["LEFT_SECOND_FRAME"][9][4] = inv_num #add inventоry number
    make_ramka(wspe, col["LEFT_SECOND_FRAME"], start)
    match countslist: #add numbers list
        case 1:
            col["CENTER_FIRST_LIST_FRAME"][32][4] = "1" #add numbers list
        case _:
            col["CENTER_FIRST_LIST_FRAME"][31][4] = "1"#add numbers list
            col["CENTER_FIRST_LIST_FRAME"][32][4] = str(countslist)#add numbers list
    col["CENTER_FIRST_LIST_FRAME"][5][4] = gbnk + gbnk_suff #add gbnk
    col["CENTER_FIRST_LIST_FRAME"][17][4] = drawby #add drawby family
    col["CENTER_FIRST_LIST_FRAME"][20][4] = namedevice + " \n Перечень элементов" # add name device
    col["CENTER_FIRST_LIST_FRAME"][25][4] = checkby #add checkby family
    col["CENTER_FIRST_LIST_FRAME"][39][4] = normby #add norm inspector family
    col["CENTER_FIRST_LIST_FRAME"][43][4] = approvby #add approved family
    make_ramka(wspe, col["CENTER_FIRST_LIST_FRAME"], start)
    f_table_r = 8 #number row frame + 1 row for first list
    name_row = type_doc + "_COL"#name key for link lists, solve type doc
    head_row = "HEAD_" + name_row#name key for head row each list, solve type doc
    k = 0
    for j in range(start, start + NROW - f_table_r - 1 , 2):
        if j == start:
            make_ramka(wspe, col[head_row], j)
        else:
            #print(data_dev[k], col[name_row])
            data_row = write_data(data_dev[k], col[name_row])
            k += 1
            make_ramka(wspe, data_row, j+1)
    s_table_r = 4 #number row frame
    for i in range(countslist-1):
        new_start_row = start+NROW+NROW*i
        make_ramka(wspe, col["LEFT_SECOND_FRAME"], new_start_row)
        col["CENTER_SECOND_LIST_FRAME"][17][4] = i + 2 #add numbers list
        col["CENTER_SECOND_LIST_FRAME"][15][4] = gbnk + gbnk_suff #add gbnk
        make_ramka(wspe, col["CENTER_SECOND_LIST_FRAME"], new_start_row)
        for j in range(new_start_row, new_start_row +NROW -s_table_r - 1 , 2):
            if j == new_start_row:
                make_ramka(wspe, col[head_row], j)
            else:
                data_row = write_data(data_dev[k], col[name_row])
                k += 1
                make_ramka(wspe, data_row, j+1)

def add_text_cell(ws:Workbook, celldata:list):
    '''additional text data'''
    r1, c1, r2, c2, cell_text, cell_rotation, cell_wrap, cell_font_size, cell_align = celldata
    text_font = 'PT Mono'
    #text_font = 'Courier New'
    cell = ws.cell(row=r1, column=c1)
    cell.font = Font(name=text_font, size=cell_font_size)
    cell.alignment = Alignment(horizontal=cell_align, vertical='center',  wrap_text=cell_wrap,
        text_rotation=cell_rotation)
    cell.value = cell_text

def write_pe_data(data_pe:list, name_row:list):
    '''write data pe in table'''
    r_name_row = name_row
    r_name_row[0][4] = data_pe[0]
    r_name_row[1][4] = data_pe[1]
    r_name_row[2][4] = data_pe[2]
    r_name_row[3][4] = data_pe[3]
    if (data_pe[0] == '') and (data_pe[2] == '') and (data_pe[3] == ''):
        r_name_row[1][8] = 'center'
    else:
        r_name_row[1][8] = 'left'
    return r_name_row

def write_sp_data(data_sp:list, name_row:list):
    '''write data sp in table'''
    r_name_row = name_row
    r_name_row[5][4] = data_sp[0]
    r_name_row[6][4] = data_sp[1]
    return r_name_row

def write_tc_data(data_tc:list, name_row:list):
    '''write data tc in table'''
    r_name_row = name_row
    r_name_row[0][4] = data_tc[0]
    r_name_row[1][4] = data_tc[1]
    r_name_row[2][4] = data_tc[2]
    r_name_row[3][4] = data_tc[3]
    r_name_row[4][4] = data_tc[4]
    return r_name_row

def create_workbook(gbnk:str, type_file:str, countlist:int, namefile, data_dev:list):
    '''create workbook and make frame'''
    wb = Workbook()
    make_document(wb, gbnk, type_file, countlist, data_dev)
    wb.save(namefile)

if __name__ == "__main__":
    HEAD_PE = ((1,3,2,6, 'Поз. обознач.', 0, 1, 12, 'center'),
            (1,7,2,13, 'Наименование', 0, 1, 12, 'center'),
            (1,14,2,15, 'Кол', 0, 1, 12, 'center'),
            (1,16,2,20, 'Примечание', 0, 1, 12, 'center'),) #sample data for table
    NAMETESTFILE = 'test.xlsx'
    COUNTSLIST = 10
    data = []
    for i in range(COUNTSLIST*25):
        data.append([i, i, i, i])
    create_workbook("ГБНК.ХХХХХХ.ХХХ", "PE", COUNTSLIST, NAMETESTFILE, data)
